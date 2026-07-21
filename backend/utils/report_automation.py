import os
import time
import calendar
import datetime
import threading
import smtplib
import json
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.extensions import db
from backend.utils.security import (
    decryptSensitiveData, mask_name, mask_email, mask_phone, mask_address, mask_city, mask_state, mask_pincode
)

# Style Constants for Excel formatting
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
DATA_FONT = Font(name="Segoe UI", size=11, color="000000")
ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Light slate
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

TITLE_FONT = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Deep Slate

def init_report_database():
    """Ensure report logs, settings, and archive tables exist in the database."""
    try:
        is_postgresql = "postgresql" in str(db.engine.url) or db.engine.name == "postgresql"
        if is_postgresql:
            # 1. Report logs table (PostgreSQL)
            db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS monthly_report_logs (
                id SERIAL PRIMARY KEY,
                report_month VARCHAR(50) NOT NULL,
                report_year INT NOT NULL,
                excel_filename VARCHAR(255) NOT NULL,
                email_status VARCHAR(50) NOT NULL,
                archive_status VARCHAR(50) NOT NULL,
                cleanup_status VARCHAR(50) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """))
            
            # 2. System settings table (PostgreSQL)
            db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS system_settings (
                setting_key VARCHAR(100) PRIMARY KEY,
                setting_value TEXT NOT NULL
            );
            """))
            
            # 3. Archive tables (PostgreSQL)
            db.session.execute(db.text("CREATE TABLE IF NOT EXISTS orders_archive (LIKE orders);"))
            db.session.execute(db.text("CREATE TABLE IF NOT EXISTS sales_archive (LIKE order_items);"))
            db.session.execute(db.text("CREATE TABLE IF NOT EXISTS buy_requests_archive (LIKE buy_requests);"))
            db.session.execute(db.text("CREATE TABLE IF NOT EXISTS support_tickets_archive (LIKE support_messages);"))
            db.session.execute(db.text("CREATE TABLE IF NOT EXISTS transactions_archive (LIKE transactions);"))
        else:
            # 1. Report logs table (MySQL/Standard)
            db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS monthly_report_logs (
                id INT AUTO_INCREMENT PRIMARY KEY,
                report_month VARCHAR(50) NOT NULL,
                report_year INT NOT NULL,
                excel_filename VARCHAR(255) NOT NULL,
                email_status VARCHAR(50) NOT NULL,
                archive_status VARCHAR(50) NOT NULL,
                cleanup_status VARCHAR(50) NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB;
            """))
            
            # 2. System settings table (MySQL/Standard)
            db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS system_settings (
                setting_key VARCHAR(100) PRIMARY KEY,
                setting_value TEXT NOT NULL
            ) ENGINE=InnoDB;
            """))
            
            # 3. Archive tables (MySQL/Standard)
            db.session.execute(db.text("CREATE TABLE IF NOT EXISTS orders_archive LIKE orders;"))
            db.session.execute(db.text("CREATE TABLE IF NOT EXISTS sales_archive LIKE order_items;"))
            db.session.execute(db.text("CREATE TABLE IF NOT EXISTS buy_requests_archive LIKE buy_requests;"))
            db.session.execute(db.text("CREATE TABLE IF NOT EXISTS support_tickets_archive LIKE support_messages;"))
            db.session.execute(db.text("CREATE TABLE IF NOT EXISTS transactions_archive LIKE transactions;"))
        
        db.session.commit()
        print("[REPORT AUTOMATION] Database tables verified/created successfully.")
        
        # Ensure email_status column is TEXT to store detailed exceptions
        try:
            if is_postgresql:
                db.session.execute(db.text("ALTER TABLE monthly_report_logs ALTER COLUMN email_status TYPE TEXT;"))
            else:
                db.session.execute(db.text("ALTER TABLE monthly_report_logs MODIFY COLUMN email_status TEXT NOT NULL;"))
            db.session.commit()
            print("[REPORT AUTOMATION] Altered email_status column to TEXT successfully.")
        except Exception as alter_err:
            db.session.rollback()
            print(f"[REPORT AUTOMATION] Alter table email_status to TEXT failed/ignored: {alter_err}")
            
    except Exception as e:
        db.session.rollback()
        print(f"[REPORT AUTOMATION ERROR] Failed to initialize tables: {e}")


def seed_default_settings():
    """Seed system SMTP and Owner settings from env if not already configured in DB."""
    try:
        res = db.session.execute(db.text("SELECT COUNT(*) FROM system_settings")).scalar()
        if res == 0:
            owner_email = os.environ.get("EMAIL_ADDRESS") or "irshadkhatola2@gmail.com"
            smtp_email = os.environ.get("MAIL_USERNAME") or os.environ.get("EMAIL_ADDRESS") or "SSJewellery2@gmail.com"
            smtp_password = os.environ.get("MAIL_PASSWORD") or os.environ.get("EMAIL_APP_PASSWORD") or "gekyxzrepafcyhem"
            
            db.session.execute(db.text("""
                INSERT INTO system_settings (setting_key, setting_value) VALUES 
                ('owner_email', :owner_email),
                ('smtp_email', :smtp_email),
                ('smtp_password', :smtp_password)
            """), {
                "owner_email": owner_email,
                "smtp_email": smtp_email,
                "smtp_password": smtp_password
            })
            db.session.commit()
            print("[REPORT AUTOMATION] Default report configurations seeded.")
    except Exception as e:
        db.session.rollback()
        print(f"[REPORT AUTOMATION ERROR] Failed to seed default settings: {e}")

def query_to_df(query, params=None):
    """Execute SQL query and return a Pandas DataFrame."""
    result = db.session.execute(db.text(query), params or {})
    columns = list(result.keys())
    data = result.fetchall()
    return pd.DataFrame(data, columns=columns)

# Reporting Data Extractors
def calculate_dashboard_stats(m, y):
    stats = {}
    # Revenue (non-cancelled orders) from both active and archived orders
    rev_q = """
        SELECT SUM(total_amount) FROM (
            SELECT total_amount, order_status, created_at FROM orders
            UNION ALL
            SELECT total_amount, order_status, created_at FROM orders_archive
        ) o
        WHERE MONTH(o.created_at) = :m AND YEAR(o.created_at) = :y AND o.order_status != 'Cancelled'
    """
    stats['total_revenue'] = float(db.session.execute(db.text(rev_q), {"m": m, "y": y}).scalar() or 0.0)
    
    # Orders Count from active and archived
    orders_q = """
        SELECT COUNT(*) FROM (
            SELECT order_status, created_at FROM orders
            UNION ALL
            SELECT order_status, created_at FROM orders_archive
        ) o
        WHERE MONTH(o.created_at) = :m AND YEAR(o.created_at) = :y
    """
    stats['total_orders'] = db.session.execute(db.text(orders_q), {"m": m, "y": y}).scalar() or 0
    
    delivered_q = """
        SELECT COUNT(*) FROM (
            SELECT order_status, created_at FROM orders
            UNION ALL
            SELECT order_status, created_at FROM orders_archive
        ) o
        WHERE MONTH(o.created_at) = :m AND YEAR(o.created_at) = :y AND o.order_status = 'Delivered'
    """
    stats['delivered_orders'] = db.session.execute(db.text(delivered_q), {"m": m, "y": y}).scalar() or 0
    
    cancelled_q = """
        SELECT COUNT(*) FROM (
            SELECT order_status, created_at FROM orders
            UNION ALL
            SELECT order_status, created_at FROM orders_archive
        ) o
        WHERE MONTH(o.created_at) = :m AND YEAR(o.created_at) = :y AND o.order_status = 'Cancelled'
    """
    stats['cancelled_orders'] = db.session.execute(db.text(cancelled_q), {"m": m, "y": y}).scalar() or 0
    
    # Users Count
    stats['registered_users'] = db.session.execute(db.text("SELECT COUNT(*) FROM users WHERE is_admin = 0")).scalar() or 0
    stats['new_users_this_month'] = db.session.execute(db.text("SELECT COUNT(*) FROM users WHERE is_admin = 0 AND MONTH(created_at) = :m AND YEAR(created_at) = :y"), {"m": m, "y": y}).scalar() or 0
    
    # Products / Stock
    stats['active_products'] = db.session.execute(db.text("SELECT COUNT(*) FROM products")).scalar() or 0
    stats['low_stock_products'] = db.session.execute(db.text("SELECT COUNT(*) FROM products WHERE stock < 10")).scalar() or 0
    
    # Buy Requests & Support (active and archived)
    buy_q = """
        SELECT COUNT(*) FROM (
            SELECT created_at FROM buy_requests
            UNION ALL
            SELECT created_at FROM buy_requests_archive
        ) br
        WHERE MONTH(br.created_at) = :m AND YEAR(br.created_at) = :y
    """
    stats['buy_requests_count'] = db.session.execute(db.text(buy_q), {"m": m, "y": y}).scalar() or 0
    
    support_q = """
        SELECT COUNT(*) FROM (
            SELECT created_at FROM support_messages
            UNION ALL
            SELECT created_at FROM support_tickets_archive
        ) sm
        WHERE MONTH(sm.created_at) = :m AND YEAR(sm.created_at) = :y
    """
    stats['support_tickets_count'] = db.session.execute(db.text(support_q), {"m": m, "y": y}).scalar() or 0
    
    return stats

def get_users_report():
    query = """
    SELECT 
        u.id AS `User ID`,
        u.full_name AS `Full Name`,
        u.email AS `Email`,
        u.phone AS `Mobile Number`,
        CASE WHEN u.is_admin = 1 THEN 'Admin' ELSE 'Customer' END AS `Role`,
        da.street AS `Address`,
        da.city AS `City`,
        da.state AS `State`,
        da.pincode AS `Pincode`,
        DATE(u.created_at) AS `Registration Date`,
        TIME(u.created_at) AS `Registration Time`,
        (
            SELECT COUNT(*) FROM (
                SELECT user_id FROM orders
                UNION ALL
                SELECT user_id FROM orders_archive
            ) o_all WHERE o_all.user_id = u.id
        ) AS `Total Orders`,
        COALESCE(
            (
                SELECT SUM(o_all.total_amount) FROM (
                    SELECT user_id, total_amount, order_status FROM orders
                    UNION ALL
                    SELECT user_id, total_amount, order_status FROM orders_archive
                ) o_all WHERE o_all.user_id = u.id AND o_all.order_status = 'Delivered'
            ), 
            0.0
        ) AS `Total Spending`,
        CASE WHEN u.is_blocked = 1 THEN 'Blocked' ELSE 'Active' END AS `Account Status`,
        u.last_login AS `Last Login`,
        (
            SELECT COUNT(*) FROM (
                SELECT user_id FROM buy_requests
                UNION ALL
                SELECT user_id FROM buy_requests_archive
            ) br_all WHERE br_all.user_id = u.id
        ) AS `Buy Requests Count`,
        (
            SELECT COUNT(*) FROM (
                SELECT email FROM support_messages
                UNION ALL
                SELECT email FROM support_tickets_archive
            ) sm_all WHERE sm_all.email = u.email
        ) AS `Support Tickets Count`
    FROM users u
    LEFT JOIN delivery_addresses da ON u.id = da.user_id
    """
    df = query_to_df(query)
    if df.empty:
        return pd.DataFrame(columns=[
            "User ID", "Full Name", "Email", "Mobile Number", "Role", "Address", "City", "State", "Pincode",
            "Registration Date", "Registration Time", "Total Orders", "Total Spending", "Account Status",
            "Last Login", "Buy Requests Count", "Support Tickets Count"
        ])
    
    # Decrypt and mask PII for report exports
    df['Full Name'] = df['Full Name'].apply(lambda x: mask_name(decryptSensitiveData(x)))
    df['Email'] = df['Email'].apply(lambda x: mask_email(decryptSensitiveData(x)))
    df['Mobile Number'] = df['Mobile Number'].apply(lambda x: mask_phone(decryptSensitiveData(x)))
    df['Address'] = df['Address'].apply(lambda x: mask_address(decryptSensitiveData(x)))
    df['City'] = df['City'].apply(lambda x: mask_city(decryptSensitiveData(x)))
    df['State'] = df['State'].apply(lambda x: mask_state(decryptSensitiveData(x)))
    df['Pincode'] = df['Pincode'].apply(lambda x: mask_pincode(decryptSensitiveData(x)))
    return df

def get_orders_report(m, y):
    query = """
    SELECT 
        o.order_id AS `Order ID`,
        u.full_name AS `Customer Name`,
        u.email AS `Customer Email`,
        u.phone AS `Customer Mobile`,
        (SELECT GROUP_CONCAT(CONCAT(oi.name, ' (x', oi.quantity, ')') SEPARATOR '; ') 
         FROM (
             SELECT name, quantity, order_id FROM order_items
             UNION ALL
             SELECT name, quantity, order_id FROM sales_archive
         ) oi WHERE oi.order_id = o.id) AS `Product Details`,
        (SELECT SUM(oi.quantity) FROM (
             SELECT quantity, order_id FROM order_items
             UNION ALL
             SELECT quantity, order_id FROM sales_archive
         ) oi WHERE oi.order_id = o.id) AS `Quantity`,
        (SELECT SUM(oi.quantity * oi.price) FROM (
             SELECT quantity, price, order_id FROM order_items
             UNION ALL
             SELECT quantity, price, order_id FROM sales_archive
         ) oi WHERE oi.order_id = o.id) AS `Price`,
        COALESCE((SELECT SUM(oi.quantity * oi.price) FROM (
             SELECT quantity, price, order_id FROM order_items
             UNION ALL
             SELECT quantity, price, order_id FROM sales_archive
         ) oi WHERE oi.order_id = o.id) - o.total_amount, 0.0) AS `Discount`,
        o.total_amount AS `Final Amount`,
        t.payment_method AS `Payment Method`,
        o.order_status AS `Order Status`,
        o.created_at AS `Order Date`,
        o.delivery_date AS `Delivery Date`,
        CASE 
            WHEN o.carrier IS NULL AND o.tracking_id IS NULL THEN 'N/A'
            ELSE CONCAT(COALESCE(o.carrier, ''), ' ', COALESCE(o.tracking_id, ''))
        END AS `Tracking Details`
    FROM (
        SELECT id, order_id, user_id, total_amount, order_status, carrier, tracking_id, created_at, delivery_date FROM orders
        UNION ALL
        SELECT id, order_id, user_id, total_amount, order_status, carrier, tracking_id, created_at, delivery_date FROM orders_archive
    ) o
    LEFT JOIN users u ON o.user_id = u.id
    LEFT JOIN (
        SELECT order_id, payment_method FROM transactions
        UNION ALL
        SELECT order_id, payment_method FROM transactions_archive
    ) t ON o.id = t.order_id
    WHERE MONTH(o.created_at) = :m AND YEAR(o.created_at) = :y
    """
    df = query_to_df(query, {"m": m, "y": y})
    if df.empty:
        return pd.DataFrame(columns=[
            "Order ID", "Customer Name", "Customer Email", "Customer Mobile", "Product Details", "Quantity",
            "Price", "Discount", "Final Amount", "Payment Method", "Order Status", "Order Date", "Delivery Date", "Tracking Details"
        ])
    
    # Decrypt and mask PII for report exports
    df['Customer Name'] = df['Customer Name'].apply(lambda x: mask_name(decryptSensitiveData(x)))
    df['Customer Email'] = df['Customer Email'].apply(lambda x: mask_email(decryptSensitiveData(x)))
    df['Customer Mobile'] = df['Customer Mobile'].apply(lambda x: mask_phone(decryptSensitiveData(x)))
    return df

def get_order_items_report(m, y):
    query = """
    SELECT 
        o.order_id AS `Order ID`,
        oi.name AS `Product Name`,
        c.name AS `Category`,
        (SELECT GROUP_CONCAT(CONCAT(pv_all.attribute_name, ': ', pv_all.attribute_value) SEPARATOR ', ') 
         FROM product_variants pv_all WHERE pv_all.product_id = oi.product_id) AS `Variant`,
        (SELECT pv_color.attribute_value FROM product_variants pv_color 
         WHERE pv_color.product_id = oi.product_id AND pv_color.attribute_name = 'Color' LIMIT 1) AS `Color`,
        (SELECT pv_size.attribute_value FROM product_variants pv_size 
         WHERE pv_size.product_id = oi.product_id AND pv_size.attribute_name = 'Size' LIMIT 1) AS `Size`,
        (SELECT pv_ram.attribute_value FROM product_variants pv_ram 
         WHERE pv_ram.product_id = oi.product_id AND pv_ram.attribute_name = 'RAM' LIMIT 1) AS `RAM`,
        (SELECT pv_storage.attribute_value FROM product_variants pv_storage 
         WHERE pv_storage.product_id = oi.product_id AND pv_storage.attribute_name = 'Storage' LIMIT 1) AS `Storage`,
        (SELECT pv_weight.attribute_value FROM product_variants pv_weight 
         WHERE pv_weight.product_id = oi.product_id AND pv_weight.attribute_name = 'Weight' LIMIT 1) AS `Weight`,
        oi.quantity AS `Quantity`,
        oi.price AS `Unit Price`,
        (oi.quantity * oi.price) AS `Total Price`
    FROM (
        SELECT id, product_id, name, quantity, price, order_id FROM order_items
        UNION ALL
        SELECT id, product_id, name, quantity, price, order_id FROM sales_archive
    ) oi
    JOIN (
        SELECT id, order_id, created_at FROM orders
        UNION ALL
        SELECT id, order_id, created_at FROM orders_archive
    ) o ON oi.order_id = o.id
    LEFT JOIN products p ON oi.product_id = p.id
    LEFT JOIN categories c ON p.category_id = c.id
    WHERE MONTH(o.created_at) = :m AND YEAR(o.created_at) = :y
    """
    df = query_to_df(query, {"m": m, "y": y})
    if df.empty:
        return pd.DataFrame(columns=[
            "Order ID", "Product Name", "Category", "Variant", "Color", "Size", "RAM", "Storage", "Weight", "Quantity", "Unit Price", "Total Price"
        ])
    return df

def get_products_report():
    query = """
    SELECT 
        p.id AS `Product ID`,
        p.name AS `Product Name`,
        c.name AS `Category`,
        'SSJewellery' AS `Brand`,
        p.price AS `Price`,
        p.discount AS `Discount`,
        p.stock AS `Stock Quantity`,
        CASE WHEN p.stock > 0 THEN 'In Stock' ELSE 'Out of Stock' END AS `Status`,
        p.created_at AS `Created Date`,
        p.updated_at AS `Updated Date`,
        COALESCE(sales.total_qty, 0) AS `Total Sales`,
        COALESCE(sales.total_rev, 0.0) AS `Total Revenue`
    FROM products p
    LEFT JOIN categories c ON p.category_id = c.id
    LEFT JOIN (
        SELECT oi.product_id, SUM(oi.quantity) AS total_qty, SUM(oi.quantity * oi.price) AS total_rev
        FROM (
            SELECT product_id, quantity, price, order_id FROM order_items
            UNION ALL
            SELECT product_id, quantity, price, order_id FROM sales_archive
        ) oi
        JOIN (
            SELECT id, order_status FROM orders
            UNION ALL
            SELECT id, order_status FROM orders_archive
        ) o ON oi.order_id = o.id
        WHERE o.order_status = 'Delivered'
        GROUP BY oi.product_id
    ) sales ON p.id = sales.product_id
    """
    df = query_to_df(query)
    if df.empty:
        return pd.DataFrame(columns=[
            "Product ID", "Product Name", "Category", "Brand", "Price", "Discount", "Stock Quantity", "Status", "Created Date", "Updated Date", "Total Sales", "Total Revenue"
        ])
    return df

def get_inventory_report():
    query = """
    SELECT 
        p.id AS `Product ID`,
        p.name AS `Product Name`,
        c.name AS `Category`,
        p.stock AS `Current Stock`,
        CASE 
            WHEN p.stock = 0 THEN 'Out of Stock' 
            WHEN p.stock <= 10 THEN 'Low Stock' 
            ELSE 'In Stock' 
        END AS `Status`
    FROM products p
    LEFT JOIN categories c ON p.category_id = c.id
    """
    df = query_to_df(query)
    if df.empty:
        return pd.DataFrame(columns=["Product ID", "Product Name", "Category", "Current Stock", "Status"])
    return df

def get_low_stock_report():
    query = """
    SELECT p.id AS `Product ID`, p.name AS `Product Name`, c.name AS `Category`, p.stock AS `Current Stock`
    FROM products p
    LEFT JOIN categories c ON p.category_id = c.id
    WHERE p.stock <= 10
    """
    df = query_to_df(query)
    if df.empty:
        return pd.DataFrame(columns=["Product ID", "Product Name", "Category", "Current Stock"])
    return df

def get_buy_requests_report(m, y):
    query = """
    SELECT 
        br.id AS `Request ID`,
        u.full_name AS `Customer Name`,
        COALESCE(br.product_name, p.name) AS `Product Name`,
        br.quantity AS `Requested Quantity`,
        br.selected_variant AS `Variant Details`,
        br.status AS `Request Status`,
        br.admin_note AS `Admin Response`,
        br.expected_availability_date AS `Expected Availability Date`,
        CASE WHEN br.status = 'Confirmed' THEN 'Confirmed' ELSE 'Not Confirmed' END AS `Confirmation Status`
    FROM (
        SELECT id, user_id, product_id, product_name, quantity, selected_variant, status, admin_note, expected_availability_date, created_at FROM buy_requests
        UNION ALL
        SELECT id, user_id, product_id, product_name, quantity, selected_variant, status, admin_note, expected_availability_date, created_at FROM buy_requests_archive
    ) br
    LEFT JOIN users u ON br.user_id = u.id
    LEFT JOIN products p ON br.product_id = p.id
    WHERE MONTH(br.created_at) = :m AND YEAR(br.created_at) = :y
    """
    df = query_to_df(query, {"m": m, "y": y})
    if df.empty:
        return pd.DataFrame(columns=[
            "Request ID", "Customer Name", "Product Name", "Requested Quantity", "Variant Details", "Request Status", "Admin Response", "Expected Availability Date", "Confirmation Status"
        ])
    
    # Decrypt and mask PII for report exports
    df['Customer Name'] = df['Customer Name'].apply(lambda x: mask_name(decryptSensitiveData(x)))
    return df

def get_support_tickets_report(m, y):
    query = """
    SELECT 
        sm.id AS `Ticket ID`,
        sm.name AS `Customer Name`,
        sm.email AS `Email`,
        CASE 
            WHEN LENGTH(sm.message) > 50 THEN CONCAT(LEFT(sm.message, 47), '...') 
            ELSE sm.message 
        END AS `Subject`,
        sm.message AS `Message`,
        'Medium' AS `Priority`,
        sm.status AS `Status`,
        sm.created_at AS `Created Date`,
        NULL AS `Resolved Date`
    FROM (
        SELECT id, name, email, message, status, created_at FROM support_messages
        UNION ALL
        SELECT id, name, email, message, status, created_at FROM support_tickets_archive
    ) sm
    WHERE MONTH(sm.created_at) = :m AND YEAR(sm.created_at) = :y
    """
    df = query_to_df(query, {"m": m, "y": y})
    if df.empty:
        return pd.DataFrame(columns=[
            "Ticket ID", "Customer Name", "Email", "Subject", "Message", "Priority", "Status", "Created Date", "Resolved Date"
        ])
    
    # Decrypt and mask PII for report exports
    df['Customer Name'] = df['Customer Name'].apply(lambda x: mask_name(decryptSensitiveData(x)))
    df['Email'] = df['Email'].apply(lambda x: mask_email(decryptSensitiveData(x)))
    return df

def get_transactions_report(m, y):
    query = """
    SELECT 
        t.transaction_id AS `Transaction ID`,
        u.full_name AS `Customer`,
        o.order_id AS `Order ID`,
        t.amount AS `Amount`,
        t.payment_method AS `Payment Method`,
        t.status AS `Payment Status`,
        DATE(t.created_at) AS `Date`,
        TIME(t.created_at) AS `Time`
    FROM (
        SELECT transaction_id, order_id, amount, payment_method, status, created_at FROM transactions
        UNION ALL
        SELECT transaction_id, order_id, amount, payment_method, status, created_at FROM transactions_archive
    ) t
    LEFT JOIN (
        SELECT id, order_id, user_id FROM orders
        UNION ALL
        SELECT id, order_id, user_id FROM orders_archive
    ) o ON t.order_id = o.id
    LEFT JOIN users u ON o.user_id = u.id
    WHERE MONTH(t.created_at) = :m AND YEAR(t.created_at) = :y
    """
    df = query_to_df(query, {"m": m, "y": y})
    if df.empty:
        return pd.DataFrame(columns=["Transaction ID", "Customer", "Order ID", "Amount", "Payment Method", "Payment Status", "Date", "Time"])
    
    # Decrypt and mask PII for report exports
    df['Customer'] = df['Customer'].apply(lambda x: mask_name(decryptSensitiveData(x)))
    return df

def write_dashboard_sheet(writer, stats, month_name, year):
    workbook = writer.book
    sheet = workbook.create_sheet("Dashboard Summary", 0)
    sheet.views.sheetView[0].showGridLines = True
    
    # Header Banner
    sheet.merge_cells("A1:G2")
    title_cell = sheet["A1"]
    title_cell.value = f"SSJewellery - MONTHLY BUSINESS REPORT ({month_name.upper()} {year})"
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row in range(1, 3):
        for col in range(1, 8):
            sheet.cell(row=row, column=col).border = THIN_BORDER
            sheet.cell(row=row, column=col).fill = TITLE_FILL
            
    # KPI Card drawer
    def draw_kpi_card(start_row, start_col, label, value, bg_hex):
        fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
        sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row+1, end_column=start_col+1)
        
        lbl_cell = sheet.cell(row=start_row, column=start_col)
        lbl_cell.value = f"{label}\n\n{value}"
        lbl_cell.font = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for r in range(start_row, start_row+2):
            for c in range(start_col, start_col+2):
                cell = sheet.cell(row=r, column=c)
                cell.fill = fill
                cell.border = THIN_BORDER

    # KPI Layout
    draw_kpi_card(4, 1, "TOTAL REVENUE", f"₹{stats['total_revenue']:,.2f}", "DCFCE7")
    draw_kpi_card(4, 3, "TOTAL ORDERS", f"{stats['total_orders']}", "DBEAFE")
    draw_kpi_card(4, 5, "DELIVERED ORDERS", f"{stats['delivered_orders']}", "D1FAE5")
    
    draw_kpi_card(7, 1, "CANCELLED ORDERS", f"{stats['cancelled_orders']}", "FEE2E2")
    draw_kpi_card(7, 3, "REGISTERED USERS", f"{stats['registered_users']}", "F3E8FF")
    draw_kpi_card(7, 5, "NEW USERS (THIS MONTH)", f"{stats['new_users_this_month']}", "E0F2FE")

    draw_kpi_card(10, 1, "ACTIVE PRODUCTS", f"{stats['active_products']}", "FEF9C3")
    draw_kpi_card(10, 3, "LOW STOCK PRODUCTS", f"{stats['low_stock_products']}", "FFEDD5")
    draw_kpi_card(10, 5, "BUY REQUESTS", f"{stats['buy_requests_count']}", "F1F5F9")
    draw_kpi_card(10, 7, "SUPPORT TICKETS", f"{stats['support_tickets_count']}", "E2E8F0")

    # Set row dimensions
    for r in [4, 5, 7, 8, 10, 11]:
        sheet.row_dimensions[r].height = 22
    for col in range(1, 9):
        sheet.column_dimensions[get_column_letter(col)].width = 16

def style_table_range(ws, start_row, start_col, df, title=None):
    current_row = start_row
    num_cols = len(df.columns)
    
    if title:
        # Title banner
        ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=start_col + num_cols - 1)
        title_cell = ws.cell(row=current_row, column=start_col)
        title_cell.value = title
        title_cell.font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid") # Slate-700
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Apply borders/fills to all merged cells in title row
        for c in range(start_col, start_col + num_cols):
            cell = ws.cell(row=current_row, column=c)
            cell.border = THIN_BORDER
            cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
            
        current_row += 1
        
    # Write Headers
    for c_idx, col_name in enumerate(df.columns):
        col_cell = ws.cell(row=current_row, column=start_col + c_idx)
        col_cell.value = col_name
        col_cell.font = HEADER_FONT
        col_cell.fill = HEADER_FILL
        col_cell.alignment = Alignment(horizontal="center", vertical="center")
        col_cell.border = THIN_BORDER
        
    current_row += 1
    
    # Write Data
    for r_idx, row in df.iterrows():
        fill = ZEBRA_FILL if r_idx % 2 == 0 else WHITE_FILL
        for c_idx, val in enumerate(row):
            cell = ws.cell(row=current_row, column=start_col + c_idx)
            cell.value = val
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            
            # Format based on header name
            col_name = str(df.columns[c_idx]).lower()
            if any(x in col_name for x in ["amount", "price", "revenue", "spending", "spent"]):
                try:
                    if val is not None:
                        cell.value = float(val)
                except ValueError:
                    pass
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif any(x in col_name for x in ["quantity", "qty", "stock", "count", "id"]):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif any(x in col_name for x in ["date", "time", "created", "updated"]):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1
        
    return current_row

def style_data_sheet(ws):
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    
    ws.row_dimensions[1].height = 28
    
    # Data Rows
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 20
        fill = ZEBRA_FILL if row % 2 == 0 else WHITE_FILL
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            
            header = str(ws.cell(row=1, column=col).value or "").lower()
            val = cell.value
            
            # Formatting
            if any(x in header for x in ["amount", "price", "revenue", "value", "spending", "spent", "discount"]):
                try:
                    if val is not None:
                        cell.value = float(val)
                except Exception:
                    pass
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif any(x in header for x in ["quantity", "qty", "stock", "count", "id", "code"]):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif any(x in header for x in ["date", "created_at", "updated_at", "time"]):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Auto-fit
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            if val is not None:
                val_str = str(val)
                if cell.number_format == '₹#,##0.00' and isinstance(val, (int, float)):
                    val_str = f"₹{val:,.2f}"
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def write_inventory_sheet(writer):
    workbook = writer.book
    ws = workbook.create_sheet("Inventory Report")
    ws.views.sheetView[0].showGridLines = True
    
    # Fetch Data
    current_stock_df = get_inventory_report()
    low_stock_df = get_low_stock_report()
    out_of_stock_df = current_stock_df[current_stock_df["Status"] == "Out of Stock"]
    
    stock_history_df = query_to_df("""
        SELECT 
            sh.id AS `History ID`,
            sh.created_at AS `Timestamp`,
            p.name AS `Product Name`,
            sh.change_type AS `Change Type`,
            sh.change_amount AS `Change Amount`,
            sh.old_stock AS `Old Stock`,
            sh.new_stock AS `New Stock`
        FROM stock_history sh
        LEFT JOIN products p ON sh.product_id = p.id
        ORDER BY sh.created_at DESC
        LIMIT 50
    """)
    
    # Handle empty dfs
    if current_stock_df.empty:
        current_stock_df = pd.DataFrame([["No current stock records found"] + [""] * 4], columns=current_stock_df.columns)
    if low_stock_df.empty:
        low_stock_df = pd.DataFrame([["No low stock alerts"] + [""] * 3], columns=low_stock_df.columns)
    if out_of_stock_df.empty:
        out_of_stock_df = pd.DataFrame([["No products out of stock"] + [""] * 4], columns=current_stock_df.columns)
    if stock_history_df.empty:
        stock_history_df = pd.DataFrame([["No inventory movements logged"] + [""] * 6], columns=stock_history_df.columns)
        
    row = 2
    row = style_table_range(ws, row, 1, current_stock_df, "Current Inventory Summary")
    row += 2
    row = style_table_range(ws, row, 1, low_stock_df, "Low Stock Items Alert")
    row += 2
    row = style_table_range(ws, row, 1, out_of_stock_df, "Out of Stock Items Alert")
    row += 2
    row = style_table_range(ws, row, 1, stock_history_df, "Inventory Movement History (Last 50 Changes)")
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def write_revenue_analytics_sheet(writer, m, y):
    workbook = writer.book
    ws = workbook.create_sheet("Revenue Analytics")
    ws.views.sheetView[0].showGridLines = True
    
    # Queries combining active and archived tables
    monthly_rev_df = query_to_df("""
        SELECT 
            DATE_FORMAT(o.created_at, '%Y-%m') AS `Month`, 
            SUM(o.total_amount) AS `Total Revenue`,
            COUNT(o.id) AS `Total Orders`
        FROM (
            SELECT id, total_amount, order_status, created_at FROM orders
            UNION ALL
            SELECT id, total_amount, order_status, created_at FROM orders_archive
        ) o
        WHERE o.order_status = 'Delivered'
        GROUP BY DATE_FORMAT(o.created_at, '%Y-%m')
        ORDER BY `Month` DESC
        LIMIT 12
    """)
    
    daily_rev_df = query_to_df("""
        SELECT 
            DATE(o.created_at) AS `Date`, 
            SUM(o.total_amount) AS `Revenue`,
            COUNT(o.id) AS `Orders`
        FROM (
            SELECT id, total_amount, order_status, created_at FROM orders
            UNION ALL
            SELECT id, total_amount, order_status, created_at FROM orders_archive
        ) o
        WHERE o.order_status = 'Delivered' 
          AND MONTH(o.created_at) = :m 
          AND YEAR(o.created_at) = :y
        GROUP BY DATE(o.created_at)
        ORDER BY `Date` ASC
    """, {"m": m, "y": y})
    
    best_selling_df = query_to_df("""
        SELECT 
            p.name AS `Product Name`,
            SUM(oi.quantity) AS `Quantity Sold`,
            SUM(oi.quantity * oi.price) AS `Total Revenue`
        FROM (
            SELECT product_id, quantity, price, order_id FROM order_items
            UNION ALL
            SELECT product_id, quantity, price, order_id FROM sales_archive
        ) oi
        JOIN (
            SELECT id, order_status FROM orders
            UNION ALL
            SELECT id, order_status FROM orders_archive
        ) o ON oi.order_id = o.id
        JOIN products p ON oi.product_id = p.id
        WHERE o.order_status = 'Delivered'
        GROUP BY oi.product_id, p.name
        ORDER BY `Quantity Sold` DESC
        LIMIT 10
    """)
    
    best_categories_df = query_to_df("""
        SELECT 
            c.name AS `Category Name`,
            SUM(oi.quantity) AS `Quantity Sold`,
            SUM(oi.quantity * oi.price) AS `Total Revenue`
        FROM (
            SELECT product_id, quantity, price, order_id FROM order_items
            UNION ALL
            SELECT product_id, quantity, price, order_id FROM sales_archive
        ) oi
        JOIN (
            SELECT id, order_status FROM orders
            UNION ALL
            SELECT id, order_status FROM orders_archive
        ) o ON oi.order_id = o.id
        JOIN products p ON oi.product_id = p.id
        JOIN categories c ON p.category_id = c.id
        WHERE o.order_status = 'Delivered'
        GROUP BY c.id, c.name
        ORDER BY `Total Revenue` DESC
    """)
    
    high_spenders_df = query_to_df("""
        SELECT 
            u.full_name AS `Customer Name`,
            u.email AS `Email`,
            COUNT(o.id) AS `Orders Placed`,
            SUM(o.total_amount) AS `Total Spent`
        FROM (
            SELECT id, user_id, total_amount, order_status FROM orders
            UNION ALL
            SELECT id, user_id, total_amount, order_status FROM orders_archive
        ) o
        JOIN users u ON o.user_id = u.id
        WHERE o.order_status = 'Delivered'
        GROUP BY o.user_id, u.full_name, u.email
        ORDER BY `Total Spent` DESC
        LIMIT 10
    """)
    
    # Format empty ones
    if monthly_rev_df.empty:
        monthly_rev_df = pd.DataFrame([["No monthly analytics available", "", ""]], columns=["Month", "Total Revenue", "Total Orders"])
    if daily_rev_df.empty:
        daily_rev_df = pd.DataFrame([["No daily data for this month", "", ""]], columns=["Date", "Revenue", "Orders"])
    if best_selling_df.empty:
        best_selling_df = pd.DataFrame([["No best selling items", "", ""]], columns=["Product Name", "Quantity Sold", "Total Revenue"])
    if best_categories_df.empty:
        best_categories_df = pd.DataFrame([["No category data", "", ""]], columns=["Category Name", "Quantity Sold", "Total Revenue"])
    if high_spenders_df.empty:
        high_spenders_df = pd.DataFrame([["No high spending customers", "", "", ""]], columns=["Customer Name", "Email", "Orders Placed", "Total Spent"])
            
    row = 2
    row = style_table_range(ws, row, 1, monthly_rev_df, "Monthly Revenue Summary (Last 12 Months)")
    row += 2
    row = style_table_range(ws, row, 1, daily_rev_df, f"Daily Revenue for Selected Month ({calendar.month_name[m]} {y})")
    row += 2
    row = style_table_range(ws, row, 1, best_selling_df, "Top 10 Best Selling Products")
    row += 2
    row = style_table_range(ws, row, 1, best_categories_df, "Product Categories by Revenue")
    row += 2
    row = style_table_range(ws, row, 1, high_spenders_df, "Top 10 Highest Spending Customers")
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def generate_report_file(m, y):
    """Generate Excel report and save to disk. Returns path."""
    month_name = calendar.month_name[m]
    reports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
    os.makedirs(reports_dir, exist_ok=True)
    
    filename = f"SSJewellery_Monthly_Report_{month_name}_{y}.xlsx"
    file_path = os.path.join(reports_dir, filename)
    
    # 1. Fetch dataframes
    users_df = get_users_report()
    orders_df = get_orders_report(m, y)
    order_items_df = get_order_items_report(m, y)
    products_df = get_products_report()
    low_stock_df = get_low_stock_report()
    buy_requests_df = get_buy_requests_report(m, y)
    support_tickets_df = get_support_tickets_report(m, y)
    transactions_df = get_transactions_report(m, y)
    
    # 2. Write to Excel
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        # We will write dashboard first manually
        stats = calculate_dashboard_stats(m, y)
        write_dashboard_sheet(writer, stats, month_name, y)
        
        # Write standard data sheets
        users_df.to_excel(writer, sheet_name="Users Report", index=False)
        orders_df.to_excel(writer, sheet_name="Orders Report", index=False)
        order_items_df.to_excel(writer, sheet_name="Order Items Report", index=False)
        products_df.to_excel(writer, sheet_name="Products Report", index=False)
        
        # Custom Inventory Report sheet
        write_inventory_sheet(writer)
        
        # Backward-compatible Low Stock sheet
        low_stock_df.to_excel(writer, sheet_name="Low Stock Products", index=False)
        
        buy_requests_df.to_excel(writer, sheet_name="Buy Requests Report", index=False)
        support_tickets_df.to_excel(writer, sheet_name="Support Tickets Report", index=False)
        transactions_df.to_excel(writer, sheet_name="Transactions Report", index=False)
        
        # Custom Revenue Analytics sheet
        write_revenue_analytics_sheet(writer, m, y)
        
        # Style all data sheets (except the custom ones which style themselves via style_table_range)
        workbook = writer.book
        custom_sheets = ["Dashboard Summary", "Inventory Report", "Revenue Analytics"]
        for name in workbook.sheetnames:
            if name not in custom_sheets:
                style_data_sheet(workbook[name])
                
    return file_path, filename

# Settings Reader Helper
def get_report_settings():
    settings = {
        "owner_email": "",
        "smtp_email": "",
        "smtp_password": ""
    }
    try:
        rows = db.session.execute(db.text("SELECT setting_key, setting_value FROM system_settings")).fetchall()
        for r in rows:
            if r[0] in settings:
                settings[r[0]] = r[1]
    except Exception as e:
        print(f"Error fetching system settings: {e}")
        
    # Fallback to env
    if not settings["owner_email"]:
        settings["owner_email"] = os.environ.get("EMAIL_ADDRESS") or "irshadkhatola2@gmail.com"
    if not settings["smtp_email"]:
        settings["smtp_email"] = os.environ.get("MAIL_USERNAME") or os.environ.get("EMAIL_ADDRESS") or "SSJewellery2@gmail.com"
    if not settings["smtp_password"]:
        settings["smtp_password"] = os.environ.get("MAIL_PASSWORD") or os.environ.get("EMAIL_APP_PASSWORD") or "gekyxzrepafcyhem"
        
    return settings

# Email Sender
def send_report_email(file_path, filename, month_name, year):
    settings = get_report_settings()
    owner_email = settings["owner_email"]
    smtp_email = settings["smtp_email"]
    smtp_password = settings["smtp_password"]
    
    # Validate credentials
    missing_fields = []
    if not owner_email:
        missing_fields.append("Owner Email")
    if not smtp_email:
        missing_fields.append("SMTP Email")
    if not smtp_password:
        missing_fields.append("SMTP App Password")
    if missing_fields:
        raise ValueError(f"SMTP configuration settings are missing: {', '.join(missing_fields)}")
        
    # Validate attachment
    if not file_path or not os.path.exists(file_path):
        raise FileNotFoundError(f"Attachment Not Found: {file_path}")
    if os.path.getsize(file_path) == 0:
        raise ValueError(f"Attachment Not Found: File size is 0 bytes: {file_path}")
        
    msg = MIMEMultipart()
    msg['From'] = smtp_email
    msg['To'] = owner_email
    msg['Subject'] = f"SSJewellery Monthly Business Report - {month_name} {year}"
    
    body = f"""Hello Owner,
 
Please find attached the automated SSJewellery Monthly Business Report for {month_name} {year}.
 
This email and report were generated automatically.
 
Warm regards,
SSJewellery Automation System"""
    msg.attach(MIMEText(body, 'plain'))
    
    # Attachment
    with open(file_path, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename= {filename}")
        msg.attach(part)
        
    # Send with retries (3 retries = total 4 attempts)
    attempts = 4
    last_exception = None
    
    for attempt in range(attempts):
        try:
            with smtplib.SMTP("smtp.gmail.com", 587, timeout=30) as server:
                server.starttls()
                server.login(smtp_email, smtp_password)
                server.send_message(msg)
            # Success! Gmail confirms delivery request
            return
        except smtplib.SMTPAuthenticationError as e:
            last_exception = Exception(f"Authentication Failed / Invalid App Password: {e}")
        except (smtplib.SMTPConnectError, TimeoutError) as e:
            last_exception = Exception(f"Connection Timeout: {e}")
        except smtplib.SMTPException as e:
            last_exception = Exception(f"SMTP Blocked: {e}")
        except Exception as e:
            last_exception = Exception(f"SMTP Error: {e}")
            
        if attempt < attempts - 1:
            print(f"[REPORT AUTOMATION] Email dispatch failed. Retrying in 30 seconds (Attempt {attempt + 1}/{attempts - 1})... Error: {last_exception}")
            time.sleep(30)
            
    raise last_exception

# Archive and Cleanup core logic
def archive_and_cleanup_data(m, y):
    """
    Archives and deletes completed target month data in database.
    If fails, it raises an exception to trigger transaction rollback.
    """
    # 1. Fetch completed orders IDs
    orders_q = """
    SELECT id FROM orders 
    WHERE MONTH(created_at) = :m AND YEAR(created_at) = :y 
      AND order_status IN ('Delivered', 'Cancelled')
    """
    order_ids = [r[0] for r in db.session.execute(db.text(orders_q), {"m": m, "y": y}).fetchall()]
    
    # 2. Fetch completed buy requests IDs (not pending)
    buy_q = """
    SELECT id FROM buy_requests 
    WHERE MONTH(created_at) = :m AND YEAR(created_at) = :y 
      AND status != 'Pending'
    """
    buy_ids = [r[0] for r in db.session.execute(db.text(buy_q), {"m": m, "y": y}).fetchall()]
    
    # 3. Fetch completed support tickets IDs (not pending)
    support_q = """
    SELECT id FROM support_messages 
    WHERE MONTH(created_at) = :m AND YEAR(created_at) = :y 
      AND status != 'Pending'
    """
    support_ids = [r[0] for r in db.session.execute(db.text(support_q), {"m": m, "y": y}).fetchall()]
    
    # Run archiving and cleanup within a single transaction
    try:
        # Move to archive tables
        if order_ids:
            ids_str = ",".join(map(str, order_ids))
            db.session.execute(db.text(f"INSERT IGNORE INTO orders_archive SELECT * FROM orders WHERE id IN ({ids_str})"))
            db.session.execute(db.text(f"INSERT IGNORE INTO sales_archive SELECT * FROM order_items WHERE order_id IN ({ids_str})"))
            db.session.execute(db.text(f"INSERT IGNORE INTO transactions_archive SELECT * FROM transactions WHERE order_id IN ({ids_str})"))
            
        if buy_ids:
            ids_str = ",".join(map(str, buy_ids))
            db.session.execute(db.text(f"INSERT IGNORE INTO buy_requests_archive SELECT * FROM buy_requests WHERE id IN ({ids_str})"))
            
        if support_ids:
            ids_str = ",".join(map(str, support_ids))
            db.session.execute(db.text(f"INSERT IGNORE INTO support_tickets_archive SELECT * FROM support_messages WHERE id IN ({ids_str})"))
            
        # Delete from active tables
        if order_ids:
            ids_str = ",".join(map(str, order_ids))
            db.session.execute(db.text(f"DELETE FROM transactions WHERE order_id IN ({ids_str})"))
            db.session.execute(db.text(f"DELETE FROM order_items WHERE order_id IN ({ids_str})"))
            db.session.execute(db.text(f"DELETE FROM orders WHERE id IN ({ids_str})"))
            
        if buy_ids:
            ids_str = ",".join(map(str, buy_ids))
            db.session.execute(db.text(f"DELETE FROM buy_requests WHERE id IN ({ids_str})"))
            
        if support_ids:
            ids_str = ",".join(map(str, support_ids))
            db.session.execute(db.text(f"DELETE FROM support_messages WHERE id IN ({ids_str})"))
            
        db.session.commit()
    except Exception as err:
        db.session.rollback()
        raise RuntimeError(f"Archive and cleanup failed: {err}")

# Central Runner Function
def run_monthly_report_flow(m, y):
    """
    Runs the full report flow: excel gen -> email delivery -> data archive -> data cleanup.
    Ensures rollback and failure protection if any step fails.
    """
    month_name = calendar.month_name[m]
    
    excel_path = None
    excel_filename = "N/A"
    email_status = "Pending"
    archive_status = "Pending"
    cleanup_status = "Pending"
    
    # 1. Excel Generation
    try:
        excel_path, excel_filename = generate_report_file(m, y)
    except Exception as e:
        # Write failure log immediately and raise
        log_report_execution(month_name, y, "N/A", "Failed", "Pending", "Pending")
        raise RuntimeError(f"Excel Generation failed: {e}")
        
    # 2. Email Delivery
    try:
        send_report_email(excel_path, excel_filename, month_name, y)
        email_status = "Success"
    except Exception as e:
        # Store exact exception / SMTP error in logs instead of generic "Failed"
        email_status = str(e)
        log_report_execution(month_name, y, excel_filename, email_status, "Pending", "Pending")
        raise RuntimeError(f"Email delivery failed: {e}")
        
    # 3. Archive Completed Records
    try:
        archive_and_cleanup_data(m, y)
        archive_status = "Success"
        cleanup_status = "Success"
    except Exception as e:
        # If archiving or cleanup fails, we log it and raise
        # Note: archive_and_cleanup_data performs database rollback on error, keeping data safe
        archive_status = "Failed" if "Archive" in str(e) else "Success"
        cleanup_status = "Failed"
        log_report_execution(month_name, y, excel_filename, email_status, archive_status, cleanup_status)
        raise e
        
    # Log complete success
    log_report_execution(month_name, y, excel_filename, email_status, archive_status, cleanup_status)
    return excel_filename

def log_report_execution(month, year, filename, email_s, archive_s, cleanup_s):
    try:
        db.session.execute(db.text("""
            INSERT INTO monthly_report_logs (report_month, report_year, excel_filename, email_status, archive_status, cleanup_status, created_at)
            VALUES (:month, :year, :filename, :email_s, :archive_s, :cleanup_s, NOW())
        """), {
            "month": month,
            "year": year,
            "filename": filename,
            "email_s": email_s,
            "archive_s": archive_s,
            "cleanup_s": cleanup_s
        })
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"[REPORT AUTOMATION ERROR] Failed to log execution: {e}")

# Scheduler runner
def check_and_run_monthly_reports():
    """
    Checks if the report for the previous month has been successfully processed.
    If not, it executes the report flow.
    This guarantees execution on the 1st of the month (or later if server was down).
    """
    now = datetime.datetime.now()
    
    # Calculate previous calendar month
    first_of_this_month = now.replace(day=1)
    last_day_of_prev_month = first_of_this_month - datetime.timedelta(days=1)
    prev_month = last_day_of_prev_month.month
    prev_year = last_day_of_prev_month.year
    prev_month_name = calendar.month_name[prev_month]
    
    # Query database log for successful run in previous month
    try:
        log_exists = db.session.execute(db.text("""
            SELECT COUNT(*) FROM monthly_report_logs 
            WHERE report_month = :month AND report_year = :year 
              AND email_status = 'Success' AND archive_status = 'Success' AND cleanup_status = 'Success'
        """), {"month": prev_month_name, "year": prev_year}).scalar() or 0
        
        if log_exists == 0:
            print(f"[REPORT SCHEDULER] Triggering automated report flow for {prev_month_name} {prev_year}...")
            run_monthly_report_flow(prev_month, prev_year)
            print(f"[REPORT SCHEDULER] Automation completed successfully for {prev_month_name} {prev_year}.")
    except Exception as e:
        print(f"[REPORT SCHEDULER ERROR] Automation failed: {e}")

# Initialization of background thread
def start_report_scheduler(app):
    def scheduler_loop():
        # Delay startup slightly to ensure database pool is ready
        time.sleep(5)
        with app.app_context():
            print("[REPORT SCHEDULER] Starting automated report worker loop...")
            init_report_database()
            seed_default_settings()
            
            while True:
                try:
                    check_and_run_monthly_reports()
                except Exception as ex:
                    print(f"[REPORT SCHEDULER LOOP EXCEPTION] {ex}")
                # Check every hour
                time.sleep(3600)
                
    t = threading.Thread(target=scheduler_loop, daemon=True)
    t.start()
